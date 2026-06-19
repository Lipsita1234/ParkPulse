"""Reports & Export routes — CSV, Excel, PDF generation."""
import io
import csv
from datetime import datetime
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse, Response
from typing import Optional

from app.ml.data_processor import get_processor
from app.ml.risk_engine import get_risk_data
from app.ml.hotspot_engine import get_hotspots

router = APIRouter(prefix="/api/reports", tags=["Reports"])


def _filter_df(location=None, station=None, vehicle=None, risk_level=None, date_from=None, date_to=None):
    proc = get_processor()
    if proc is None or proc.df is None:
        return None
    df = proc.df.copy()

    if location:
        df = df[df["location"].str.contains(location, case=False, na=False)]
    if station and "police_station" in df.columns:
        df = df[df["police_station"].str.contains(station, case=False, na=False)]
    if vehicle and "vehicle_type" in df.columns:
        df = df[df["vehicle_type"].str.lower() == vehicle.lower()]
    if date_from and "created_date" in df.columns:
        df = df[df["created_date"] >= date_from]
    if date_to and "created_date" in df.columns:
        df = df[df["created_date"] <= date_to]

    return df


@router.get("/violations/csv")
async def export_violations_csv(
    location: Optional[str] = None,
    station: Optional[str] = None,
    vehicle: Optional[str] = None,
    limit: int = Query(5000, le=50000),
):
    df = _filter_df(location=location, station=station, vehicle=vehicle)
    if df is None:
        return Response("No data", status_code=404)

    export_cols = [c for c in ["id", "location_short", "vehicle_type", "violation", 
                                "police_station", "junction", "hour", "day_name",
                                "month_name", "risk_level"] if c in df.columns]
    df = df[export_cols].head(limit)

    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=violations_{datetime.now().strftime('%Y%m%d')}.csv"},
    )


@router.get("/violations/excel")
async def export_violations_excel(
    location: Optional[str] = None,
    station: Optional[str] = None,
    vehicle: Optional[str] = None,
    limit: int = Query(5000, le=50000),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    df = _filter_df(location=location, station=station, vehicle=vehicle)
    if df is None:
        return Response("No data", status_code=404)

    export_cols = [c for c in ["id", "location_short", "vehicle_type", "violation",
                                "police_station", "junction", "hour", "day_name", "month_name"] if c in df.columns]
    df = df[export_cols].head(limit)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Violations"

    # Header style
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx + 2, column=col_idx, value=str(val) if val else "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=violations_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get("/hotspots/csv")
async def export_hotspots_csv():
    hotspots = get_hotspots()
    if not hotspots:
        return Response("No data", status_code=404)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=hotspots[0].keys())
    writer.writeheader()
    writer.writerows(hotspots)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=hotspots.csv"},
    )


@router.get("/risk/csv")
async def export_risk_csv():
    df = get_risk_data()
    if df is None:
        return Response("No data", status_code=404)
    cols = ["location", "risk_score", "risk_level", "violation_count", "police_station"]
    output = io.StringIO()
    df[cols].to_csv(output, index=False)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=risk_analysis.csv"},
    )


@router.get("/pdf")
async def export_pdf_report():
    """Generate a professional PDF summary report."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, KeepTogether
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import Frame, PageTemplate
        from reportlab.pdfgen import canvas as rl_canvas

        proc = get_processor()
        hotspots = get_hotspots()
        risk_df = get_risk_data()

        # ── Color Palette ──────────────────────────────────────────
        NAVY       = colors.HexColor("#1E3A5F")
        NAVY_LIGHT = colors.HexColor("#2D5A8E")
        BLUE       = colors.HexColor("#3B82F6")
        RED        = colors.HexColor("#EF4444")
        ORANGE     = colors.HexColor("#F97316")
        GREEN      = colors.HexColor("#22C55E")
        GRAY_BG    = colors.HexColor("#F8FAFC")
        GRAY_LINE  = colors.HexColor("#E2E8F0")
        GRAY_TEXT  = colors.HexColor("#64748B")
        WHITE      = colors.white
        DARK       = colors.HexColor("#1E293B")

        now = datetime.now()
        doc_num = f"PP-{now.strftime('%Y%m%d-%H%M')}"
        date_str = now.strftime("%d %B %Y  ·  %H:%M IST")

        buffer = io.BytesIO()
        page_w, page_h = A4

        # ── Header / Footer drawn on every page ────────────────────
        def draw_header_footer(canvas_obj, doc_obj):
            canvas_obj.saveState()

            # Top header bar
            canvas_obj.setFillColor(NAVY)
            canvas_obj.rect(0, page_h - 42*mm, page_w, 42*mm, fill=1, stroke=0)

            # Logo / Org name
            canvas_obj.setFont("Helvetica-Bold", 16)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.drawString(18*mm, page_h - 18*mm, "ParkPulse AI")
            canvas_obj.setFont("Helvetica", 7.5)
            canvas_obj.setFillColor(colors.HexColor("#93C5FD"))
            canvas_obj.drawString(18*mm, page_h - 26*mm,
                "Smart Parking Enforcement Intelligence Platform  ·  Bengaluru Traffic Authority")

            # Doc number + date (top-right)
            canvas_obj.setFont("Helvetica-Bold", 8)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.drawRightString(page_w - 18*mm, page_h - 17*mm, f"Report No: {doc_num}")
            canvas_obj.setFont("Helvetica", 7.5)
            canvas_obj.setFillColor(colors.HexColor("#93C5FD"))
            canvas_obj.drawRightString(page_w - 18*mm, page_h - 25*mm, date_str)

            # Report title band
            canvas_obj.setFillColor(NAVY_LIGHT)
            canvas_obj.rect(0, page_h - 52*mm, page_w, 10*mm, fill=1, stroke=0)
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.drawString(18*mm, page_h - 47.5*mm,
                "ENFORCEMENT INTELLIGENCE EXECUTIVE REPORT  —  CONFIDENTIAL")

            # Bottom footer bar
            canvas_obj.setFillColor(NAVY)
            canvas_obj.rect(0, 0, page_w, 14*mm, fill=1, stroke=0)
            canvas_obj.setFont("Helvetica", 7)
            canvas_obj.setFillColor(colors.HexColor("#93C5FD"))
            canvas_obj.drawString(18*mm, 5*mm,
                f"ParkPulse AI  ·  Bengaluru Traffic Authority  ·  {doc_num}  ·  AI-generated report — for official use only")
            canvas_obj.setFont("Helvetica-Bold", 8)
            canvas_obj.setFillColor(WHITE)
            canvas_obj.drawRightString(page_w - 18*mm, 5*mm,
                f"Page {doc_obj.page}")

            canvas_obj.restoreState()

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=18*mm, rightMargin=18*mm,
            topMargin=57*mm, bottomMargin=20*mm,
            onFirstPage=draw_header_footer, onLaterPages=draw_header_footer,
        )

        # ── Paragraph Styles ───────────────────────────────────────
        styles = getSampleStyleSheet()

        def ps(name, **kw):
            return ParagraphStyle(name, **kw)

        S_SECTION = ps("Section", fontSize=11, fontName="Helvetica-Bold",
                        textColor=NAVY, spaceBefore=14, spaceAfter=6)
        S_BODY    = ps("Body", fontSize=9, fontName="Helvetica",
                        textColor=DARK, leading=14, spaceAfter=4)
        S_LABEL   = ps("Label", fontSize=7.5, fontName="Helvetica-Bold",
                        textColor=GRAY_TEXT, spaceAfter=2)
        S_SMALL   = ps("Small", fontSize=7.5, fontName="Helvetica",
                        textColor=GRAY_TEXT, leading=11)

        def section_heading(text, color=NAVY):
            return [
                Paragraph(text, S_SECTION),
                HRFlowable(width="100%", thickness=1.5, color=color, spaceAfter=8),
            ]

        story = []

        # ── 1. Summary KPI boxes ───────────────────────────────────
        story += section_heading("Executive Summary — City-Wide Statistics")

        if proc and proc.df is not None:
            df = proc.df
            total_v   = len(df)
            unique_l  = df["location"].nunique() if "location" in df.columns else "N/A"
            stations  = df["police_station"].nunique() if "police_station" in df.columns else "N/A"
            clusters  = len(hotspots)
            high_risk = len([h for h in hotspots if h.get("risk_level") == "High"])
            med_risk  = len([h for h in hotspots if h.get("risk_level") == "Medium"])

            kpi_items = [
                ("Total Violations Analysed", f"{total_v:,}"),
                ("Unique Road Locations",     f"{unique_l:,}"),
                ("Police Station Precincts",  str(stations)),
                ("DBSCAN Hotspot Clusters",   str(clusters)),
                ("High Risk Locations",       str(high_risk)),
                ("Medium Risk Locations",     str(med_risk)),
            ]
            # 3-column kpi table
            kpi_rows = []
            for i in range(0, len(kpi_items), 3):
                trio = kpi_items[i:i+3]
                labels = [Paragraph(t[0], S_LABEL) for t in trio]
                values = [Paragraph(t[1],
                            ps(f"v{i}", fontSize=18, fontName="Helvetica-Bold",
                               textColor=BLUE, spaceAfter=0))
                          for t in trio]
                # pad if last row short
                while len(labels) < 3:
                    labels.append(Paragraph("", S_LABEL))
                    values.append(Paragraph("", S_LABEL))
                kpi_rows.append(labels)
                kpi_rows.append(values)

            col_w = (page_w - 36*mm) / 3
            kpi_t = Table(kpi_rows, colWidths=[col_w]*3, repeatRows=0)
            kpi_t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), GRAY_BG),
                ("BOX",       (0,0), (-1,-1), 0.5, GRAY_LINE),
                ("INNERGRID", (0,0), (-1,-1), 0.5, GRAY_LINE),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 8),
                ("LEFTPADDING",   (0,0), (-1,-1), 10),
                ("RIGHTPADDING",  (0,0), (-1,-1), 10),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ]))
            story.append(kpi_t)
            story.append(Spacer(1, 5*mm))

        # ── 2. Top Hotspot Locations ───────────────────────────────
        if hotspots:
            story += section_heading("Top 10 High-Risk Enforcement Zones", RED)

            def risk_color(level):
                return {
                    "High": colors.HexColor("#FEE2E2"),
                    "Medium": colors.HexColor("#FFF7ED"),
                    "Low": colors.HexColor("#F0FDF4"),
                }.get(level, WHITE)

            def risk_text_color(level):
                return {"High": RED, "Medium": ORANGE, "Low": GREEN}.get(level, DARK)

            hs_header = [
                Paragraph("#",              ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
                Paragraph("Location",       ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
                Paragraph("Risk Score",     ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                Paragraph("Risk Level",     ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                Paragraph("Violations",     ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                Paragraph("Police Station", ps("h", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
            ]
            hs_data = [hs_header]
            cmds = [
                ("BACKGROUND", (0,0), (-1,0), RED),
                ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("GRID",       (0,0), (-1,-1), 0.4, GRAY_LINE),
                ("TOPPADDING",    (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LEFTPADDING",   (0,0), (-1,-1), 8),
                ("RIGHTPADDING",  (0,0), (-1,-1), 8),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ("ALIGN",         (2,0), (4,-1), "CENTER"),
            ]
            for idx, h in enumerate(hotspots[:10], 1):
                lvl = h.get("risk_level", "Low")
                row_bg = risk_color(lvl) if idx % 2 == 0 else WHITE
                hs_data.append([
                    Paragraph(str(idx), S_SMALL),
                    Paragraph(h.get("location_name", "")[:42], S_SMALL),
                    Paragraph(str(h.get("risk_score", "")),
                              ps(f"rs{idx}", fontSize=8, fontName="Helvetica-Bold",
                                 textColor=risk_text_color(lvl), alignment=TA_CENTER)),
                    Paragraph(lvl,
                              ps(f"rl{idx}", fontSize=8, fontName="Helvetica-Bold",
                                 textColor=risk_text_color(lvl), alignment=TA_CENTER)),
                    Paragraph(f"{h.get('violation_count', 0):,}", S_SMALL),
                    Paragraph(h.get("police_station", "")[:22], S_SMALL),
                ])
                cmds.append(("BACKGROUND", (0, idx), (-1, idx), row_bg))

            hs_t = Table(hs_data, colWidths=[10*mm, 62*mm, 22*mm, 22*mm, 22*mm, 37*mm])
            hs_t.setStyle(TableStyle(cmds))
            story.append(hs_t)
            story.append(Spacer(1, 5*mm))

        # ── 3. Risk Level Distribution ────────────────────────────
        if risk_df is not None and not risk_df.empty:
            story += section_heading("Risk Distribution by Location")
            if "risk_level" in risk_df.columns:
                dist = risk_df["risk_level"].value_counts()
                total_locs = len(risk_df)
                dist_header = [
                    Paragraph("Risk Category",   ps("dh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
                    Paragraph("Locations",        ps("dh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                    Paragraph("% of Total",       ps("dh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                    Paragraph("Avg. Violations",  ps("dh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER)),
                ]
                dist_data = [dist_header]
                dist_cmds = [
                    ("BACKGROUND", (0,0), (-1,0), NAVY),
                    ("FONTSIZE",   (0,0), (-1,-1), 8),
                    ("GRID",       (0,0), (-1,-1), 0.4, GRAY_LINE),
                    ("TOPPADDING",    (0,0), (-1,-1), 5),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                    ("LEFTPADDING",   (0,0), (-1,-1), 10),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 10),
                    ("ALIGN",         (1,0), (-1,-1), "CENTER"),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ]
                for i, lvl in enumerate(["High", "Medium", "Low"], 1):
                    cnt = int(dist.get(lvl, 0))
                    pct = (cnt / total_locs * 100) if total_locs else 0
                    avg_v = ""
                    if "violation_count" in risk_df.columns and "risk_level" in risk_df.columns:
                        sub = risk_df[risk_df["risk_level"] == lvl]
                        avg_v = f"{sub['violation_count'].mean():,.0f}" if not sub.empty else "—"
                    rc = risk_text_color(lvl)
                    bg = risk_color(lvl)
                    dist_data.append([
                        Paragraph(lvl, ps(f"dl{i}", fontSize=8, fontName="Helvetica-Bold", textColor=rc)),
                        Paragraph(f"{cnt:,}", S_SMALL),
                        Paragraph(f"{pct:.1f}%", S_SMALL),
                        Paragraph(avg_v, S_SMALL),
                    ])
                    dist_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))

                col_w = (page_w - 36*mm) / 4
                dist_t = Table(dist_data, colWidths=[col_w]*4)
                dist_t.setStyle(TableStyle(dist_cmds))
                story.append(dist_t)
                story.append(Spacer(1, 5*mm))

        # ── 4. Enforcement Deployment Guidance ────────────────────
        story += section_heading("Standard Enforcement Deployment Guidelines")
        deploy_data = [
            [Paragraph("Risk Level", ps("gh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
             Paragraph("Priority",   ps("gh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
             Paragraph("Officers",   ps("gh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE)),
             Paragraph("Monitoring Schedule", ps("gh", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE))],
            [Paragraph("HIGH",   ps("gl1", fontSize=8, fontName="Helvetica-Bold", textColor=RED)),
             Paragraph("IMMEDIATE / CRITICAL", S_SMALL),
             Paragraph("3–5 officers", S_SMALL),
             Paragraph("Continuous 24/7 monitoring", S_SMALL)],
            [Paragraph("MEDIUM", ps("gl2", fontSize=8, fontName="Helvetica-Bold", textColor=ORANGE)),
             Paragraph("SCHEDULED / MODERATE", S_SMALL),
             Paragraph("1–2 officers", S_SMALL),
             Paragraph("Peak hours: 07:00–10:00, 17:00–21:00", S_SMALL)],
            [Paragraph("LOW",    ps("gl3", fontSize=8, fontName="Helvetica-Bold", textColor=GREEN)),
             Paragraph("ROUTINE Surveillance", S_SMALL),
             Paragraph("Standard patrol", S_SMALL),
             Paragraph("Normal surveillance schedule", S_SMALL)],
        ]
        deploy_cmds = [
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#FEE2E2")),
            ("BACKGROUND", (0,2), (-1,2), colors.HexColor("#FFF7ED")),
            ("BACKGROUND", (0,3), (-1,3), colors.HexColor("#F0FDF4")),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("GRID",       (0,0), (-1,-1), 0.4, GRAY_LINE),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("RIGHTPADDING",  (0,0), (-1,-1), 10),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]
        deploy_t = Table(deploy_data, colWidths=[28*mm, 45*mm, 35*mm, 67*mm])
        deploy_t.setStyle(TableStyle(deploy_cmds))
        story.append(deploy_t)
        story.append(Spacer(1, 6*mm))

        # ── 5. Disclaimer ─────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY_LINE, spaceAfter=6))
        story.append(Paragraph(
            f"DISCLAIMER: This report is AI-generated by the ParkPulse Smart Enforcement Intelligence Platform "
            f"using historical parking violation datasets for Bengaluru. All metrics are derived from statistical "
            f"analysis and machine learning models. Field officers must apply professional discretion. "
            f"Report Reference: {doc_num} · Generated: {date_str}.",
            S_SMALL
        ))

        # ── Build ──────────────────────────────────────────────────
        doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=parkpulse_report_{now.strftime('%Y%m%d')}.pdf"},
        )
    except Exception as e:
        import traceback
        return Response(f"PDF generation error: {e}\n{traceback.format_exc()}", status_code=500)
